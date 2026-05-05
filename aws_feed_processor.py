"""
AWS Feed Processor v5
---------------------
Fetches the latest AWS announcements from the RSS feed,
deduplicates against a persistent master Excel file,
processes only NEW entries using Amazon Bedrock Nova Lite (cross-region inference),
extracts AWS Service and Category categorization,
and appends results along with a cost summary to the master Excel file.

Deduplication logic:
- Day 1 : master file does not exist → process all 100 entries → create master file
- Day 2+: master file exists → load processed links → skip already seen entries → process only new ones
"""

# ─────────────────────────────────────────────
# Imports
# ─────────────────────────────────────────────
import sys
import os
import re
import json
import time
import logging
from datetime import datetime

import boto3
import feedparser
from botocore.config import Config
from botocore.exceptions import ClientError
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

# ─────────────────────────────────────────────
# UTF-8 stdout — must be before logging setup
# ─────────────────────────────────────────────
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

# ─────────────────────────────────────────────
# Logging Setup
# ─────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("aws_feed_processor.log", encoding="utf-8")
    ]
)
logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────
# Configuration
# ─────────────────────────────────────────────
FEED_URL         = "https://aws.amazon.com/new/feed/"
MODEL_ID         = "apac.amazon.nova-lite-v1:0"
AWS_REGION       = "ap-south-1"
MAX_ENTRIES      = 100
MAX_TITLE_LEN    = 300
MAX_SUMMARY_LEN  = 2000
MAX_RETRIES      = 3
RATE_LIMIT_DELAY = 0.5   # seconds between successful API calls

# Persistent master file — same file reused every run
MASTER_FILE      = "aws_feed_master.xlsx"

# Nova Lite cross-region inference pricing — Asia Pacific (Mumbai) APS3 actual rates
INPUT_COST_PER_1M  = 0.071   # $0.071 per 1M input tokens
OUTPUT_COST_PER_1M = 0.284   # $0.284 per 1M output tokens

# ─────────────────────────────────────────────
# AWS Bedrock Client
# ─────────────────────────────────────────────
bedrock = boto3.client(
    "bedrock-runtime",
    region_name=AWS_REGION,
    config=Config(connect_timeout=5, read_timeout=30)
)

# ─────────────────────────────────────────────
# Helper Functions
# ─────────────────────────────────────────────
def strip_html(text: str) -> str:
    """Remove HTML tags from text."""
    return re.sub(r'<[^>]+>', ' ', text).strip()


def build_prompt(title: str, summary: str) -> str:
    """Build the prompt to send to the model with categorization."""
    return (
        f"AWS Update:\nTitle: {title}\nSummary: {summary}\n\n"
        "Provide:\n"
        "1. Short Summary (2 lines)\n"
        "2. Outcome (impact)\n"
        "3. Where to use (use case)\n"
        "4. AWS Service: <service name> | Category: <category>\n\n"
        "For Category, choose one: AI/ML, Compute, Database, Networking, Security, "
        "Storage, Analytics, Developer Tools, Management, or Other.\n"
        "Keep it concise."
    )


def validate_and_clean(title: str, summary: str) -> tuple:
    """Strip HTML, truncate, and validate title and summary."""
    title   = strip_html(title)[:MAX_TITLE_LEN].strip()
    summary = strip_html(summary)[:MAX_SUMMARY_LEN].strip()

    if not title or not summary:
        raise ValueError("Title or summary is empty after validation")

    return title, summary


def parse_llm_output(raw_output: str) -> dict:
    """
    Parse LLM output to extract summary, AWS service, and category.
    Returns: {
        "summary":  "1. Short Summary...\n2. Outcome...\n3. Where to use...",
        "service":  "Amazon EKS",
        "category": "Compute"
    }
    """
    service_match  = re.search(r'AWS Service:\s*([^|]+)',     raw_output, re.IGNORECASE)
    category_match = re.search(r'Category:\s*(.+?)(?:\n|$)', raw_output, re.IGNORECASE)

    service  = service_match.group(1).strip()  if service_match  else "Unknown"
    category = category_match.group(1).strip() if category_match else "Unknown"

    # Remove line 4 from the displayed summary
    summary = re.sub(
        r'4\.\s*AWS Service:.*?Category:.*?(?:\n|$)', '',
        raw_output,
        flags=re.IGNORECASE | re.DOTALL
    ).strip()

    return {
        "summary":  summary,
        "service":  service,
        "category": category
    }


# ─────────────────────────────────────────────
# Deduplication
# ─────────────────────────────────────────────
def load_processed_links(filepath: str) -> set:
    """
    Load all previously processed entry links from the master Excel file.
    Returns an empty set if the master file does not exist yet (Day 1).
    """
    if not os.path.exists(filepath):
        logger.info("Master file not found — fresh run, all entries will be processed")
        return set()

    wb    = load_workbook(filepath)
    ws    = wb["AWS Feed"]
    links = {row[1] for row in ws.iter_rows(min_row=2, values_only=True) if row[1]}
    logger.info("Loaded %d previously processed links from master file", len(links))
    return links


def filter_new_entries(entries: list, processed_links: set) -> list:
    """Return only entries whose link has not been processed before."""
    new_entries = [e for e in entries if e.get("link", "") not in processed_links]
    logger.info(
        "Deduplication: %d in feed | %d already processed | %d new to process",
        len(entries), len(processed_links), len(new_entries)
    )
    return new_entries


# ─────────────────────────────────────────────
# Bedrock Inference
# ─────────────────────────────────────────────
def call_bedrock(prompt: str) -> dict:
    """
    Invoke the Bedrock model and return the raw result dict.
    Retries on ThrottlingException with exponential backoff.
    """
    for attempt in range(MAX_RETRIES):
        try:
            response = bedrock.invoke_model(
                modelId=MODEL_ID,
                body=json.dumps({
                    "messages": [{"role": "user", "content": [{"text": prompt}]}],
                    "inferenceConfig": {
                        "maxTokens": 350,
                        "temperature": 0.3,
                        "stopSequences": []
                    }
                }),
                contentType="application/json",
                accept="application/json"
            )
            return json.loads(response["body"].read())

        except ClientError as e:
            error_code = e.response["Error"]["Code"]
            if error_code == "ThrottlingException" and attempt < MAX_RETRIES - 1:
                wait = 2 ** attempt
                logger.warning("Throttled. Retrying in %ds (attempt %d/%d)", wait, attempt + 1, MAX_RETRIES)
                time.sleep(wait)
                continue
            raise

    raise RuntimeError("Max retries exceeded for Bedrock call")


def process_entry(title: str, summary: str, token_counts: dict) -> dict:
    """
    Clean input, call Bedrock, update token counts, and return parsed output.
    Returns: {"summary": str, "service": str, "category": str}
    """
    title, summary = validate_and_clean(title, summary)
    prompt         = build_prompt(title, summary)
    result         = call_bedrock(prompt)

    usage = result.get("usage", {})
    token_counts["input"]  += usage.get("inputTokens", 0)
    token_counts["output"] += usage.get("outputTokens", 0)

    raw_output = result["output"]["message"]["content"][0]["text"].strip()
    raw_output = re.sub(r'\*+', '', raw_output).strip()
    return parse_llm_output(raw_output)


# ─────────────────────────────────────────────
# Excel: Load or Create Master Workbook
# ─────────────────────────────────────────────
def load_or_create_workbook(filepath: str):
    """
    Load existing master workbook if it exists, otherwise create a new one.
    Returns (workbook, feed_sheet).
    """
    if os.path.exists(filepath):
        logger.info("Loading existing master file: %s", filepath)
        wb = load_workbook(filepath)
        ws = wb["AWS Feed"]
    else:
        logger.info("Creating new master file: %s", filepath)
        wb = Workbook()
        ws = wb.active
        ws.title = "AWS Feed"
        ws.append(["Title", "Link", "Published", "LLM_Output", "AWS_Service", "Category"])
        format_header(ws)

    return wb, ws


# ─────────────────────────────────────────────
# Excel Formatting
# ─────────────────────────────────────────────
def format_header(ws) -> None:
    """Apply header formatting and fixed column widths."""
    col_widths = [60, 50, 25, 80, 30, 20]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = width

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")


# ─────────────────────────────────────────────
# Output: Excel Writer
# ─────────────────────────────────────────────
def append_feed_entries(ws, entries: list, token_counts: dict) -> int:
    """Process new entries and append results to the AWS Feed sheet. Returns error count."""
    total       = len(entries)
    error_count = 0

    for i, entry in enumerate(entries, start=1):
        title     = entry.get("title", "")
        summary   = entry.get("summary", "")
        link      = entry.get("link", "")
        published = entry.get("published", "")

        logger.info("Processing entry %d/%d: %s", i, total, title[:60])

        try:
            parsed      = process_entry(title, summary, token_counts)
            llm_output  = parsed["summary"]
            aws_service = parsed["service"]
            category    = parsed["category"]
            logger.info("Entry %d/%d -> Service: %s | Category: %s", i, total, aws_service, category)
            time.sleep(RATE_LIMIT_DELAY)  # polite delay only on success
        except Exception as e:
            llm_output  = f"Error: {str(e)}"
            aws_service = "Error"
            category    = "Error"
            error_count += 1
            logger.error("Entry %d/%d failed: %s", i, total, e)

        ws.append([title, link, published, llm_output, aws_service, category])

    return error_count


def update_cost_sheet(wb: Workbook, token_counts: dict) -> None:
    """
    Update the Cost Summary sheet.
    Replaces it if it already exists to reflect the latest run's cost.
    Warns if the sheet was missing when accumulating monthly cost.
    """
    cumulative_monthly_cost = 0.0

    if "Cost Summary" in wb.sheetnames:
        ws_existing = wb["Cost Summary"]
        for row in ws_existing.iter_rows(min_row=2, values_only=True):
            if row[0] == "Cumulative Monthly Cost (USD)":
                cumulative_monthly_cost = float(row[1] or 0)
                break
        del wb["Cost Summary"]
    else:
        logger.warning("Cost Summary sheet missing — cumulative monthly cost reset to 0.0")

    input_tokens  = token_counts["input"]
    output_tokens = token_counts["output"]
    input_cost    = (input_tokens  / 1_000_000) * INPUT_COST_PER_1M
    output_cost   = (output_tokens / 1_000_000) * OUTPUT_COST_PER_1M
    total_cost    = input_cost + output_cost
    cumulative_monthly_cost = round(cumulative_monthly_cost + total_cost, 6)

    ws = wb.create_sheet(title="Cost Summary")
    ws.append(["Metric", "Value"])
    ws.append(["Model",                         MODEL_ID])
    ws.append(["Pricing note",                  "Nova Lite APS3 rates as of 2025-07"])
    ws.append(["Run Date",                      datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["New Entries Processed",         token_counts.get("entries", 0)])
    ws.append(["Total Input Tokens",            input_tokens])
    ws.append(["Total Output Tokens",           output_tokens])
    ws.append(["Input Cost (USD)",              round(input_cost,  6)])
    ws.append(["Output Cost (USD)",             round(output_cost, 6)])
    ws.append(["Total Cost This Run (USD)",     round(total_cost,  6)])
    ws.append(["Cumulative Monthly Cost (USD)", cumulative_monthly_cost])

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 55
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    logger.info(
        "Cost Summary -> Input: %d tokens | Output: %d tokens | This run: $%.6f | Cumulative: $%.6f",
        input_tokens, output_tokens, total_cost, cumulative_monthly_cost
    )


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────
def main():
    logger.info("Starting AWS Feed Processor v5")

    # Step 1: Fetch RSS feed
    logger.info("Fetching feed: %s", FEED_URL)
    feed    = feedparser.parse(FEED_URL)
    entries = feed.entries[:MAX_ENTRIES]

    if not entries:
        logger.error("No entries returned from feed — check URL or network connection.")
        sys.exit(1)

    logger.info("Feed returned %d entries", len(entries))

    # Step 2: Load already processed links from master file
    processed_links = load_processed_links(MASTER_FILE)

    # Step 3: Filter to only new entries
    new_entries = filter_new_entries(entries, processed_links)

    if not new_entries:
        logger.info("No new entries found — nothing to process. Exiting.")
        return

    # Step 4: Load or create master workbook
    wb, ws = load_or_create_workbook(MASTER_FILE)

    # Step 5: Process new entries and append to master sheet
    token_counts = {"input": 0, "output": 0, "entries": len(new_entries)}
    error_count  = append_feed_entries(ws, new_entries, token_counts)

    # Step 6: Update cost summary sheet
    update_cost_sheet(wb, token_counts)

    # Step 7: Save master file
    wb.save(MASTER_FILE)
    logger.info(
        "Done. %d new entries appended (%d errors) -> %s",
        len(new_entries), error_count, MASTER_FILE
    )


if __name__ == "__main__":
    main()
